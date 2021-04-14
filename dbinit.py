# -*- coding: utf-8 -*-
"""
Created on Thu Mar 18 12:36:09 2021

@author: nobin
"""
import pymongo
import openpyxl
client=pymongo.MongoClient()
Medicine=client.Medical_Database
Doctor=client.Doctor_Database
Nurse=client.Nurse_Database
Stock=client.Stock_Database
Patient=client.Patient_Database
medicine_price_table=Medicine.Table
Doctor_table=Doctor.Table
Nurse_table=Nurse.Table
Patient_table=Patient.Table
Stock_table=Stock.Table
class Dbinit():
    def __init__(self):
        dbDocument=openpyxl.load_workbook("Medicines.xlsx")
        Medicine_price=dbDocument['Sheet1']
        Doctor_list=dbDocument['Sheet2']
        Patient_list=dbDocument['Sheet3']
        Nurse_list=dbDocument['Sheet4']
        Stock_list=dbDocument['Sheet5']
        medicine_price_table.drop()
        Doctor_table.drop()
        Nurse_table.drop()
        Patient_table.drop()
        Stock_table.drop()
        for rows in range(2,Medicine_price.max_row+1):
            table_dict={'Medicine Name':Medicine_price.cell(rows,1).value.upper(),'Package units':int(Medicine_price.cell(rows,3).value),'Price':float(Medicine_price.cell(rows,4).value)}
            medicine_price_table.insert_one(table_dict)
        for rows in range(2,Doctor_list.max_row+1):
            table_dict={'Doctor_id':'Dr'+str(Patient_list.cell(rows,1).value),'Name':Doctor_list.cell(rows,2).value,'Ward':Doctor_list.cell(rows,3).value}
            Doctor_table.insert_one(table_dict)
        for rows in range(2,Nurse_list.max_row+1):
            table_dict={'Nurse_id':'Nr'+str(Nurse_list.cell(rows,1).value),'Name':Nurse_list.cell(rows,2).value,'Ward':Nurse_list.cell(rows,3).value}
            Nurse_table.insert_one(table_dict)
        for rows in range(2,Patient_list.max_row+1):
            table_dict={'Patient_id':'Pt'+str(Patient_list.cell(rows,1).value),'Name':Patient_list.cell(rows,2).value,'Ward':Patient_list.cell(rows,3).value,'Prescribed Medicines':None}
            Patient_table.insert_one(table_dict)
        for rows in range(2,Stock_list.max_row+1):
            table_dict={'Medicine Name':Stock_list.cell(rows,1).value.upper(),'Units Remaining':int(Stock_list.cell(rows,3).value)}
            Stock_table.insert_one(table_dict)
Example=Dbinit()
for document in Stock_table.find():
    print(document)
    