import pymongo
import datetime
from selenium import webdriver
import openpyxl
client=pymongo.MongoClient()
idDocument=openpyxl.load_workbook("id's.xlsx")
ids=idDocument['Sheet1']
htmlids=client.id_database
id_table=htmlids.Table
id_table.drop()
for rows in range(2,ids.max_row+1):
            table_dict={'Medicine Name':ids.cell(rows,1).value.upper(),'id':ids.cell(rows,2).value}
            id_table.insert_one(table_dict)
Medicine=client.Medical_Database
Stock=client.Stock_Database
Patient=client.Patient_Database
medicine_price_table=Medicine.Table
Patient_table=Patient.Table
Stock_table=Stock.Table

while True:
    price=0
    mlist=[]
    if datetime.datetime.now().strftime("%H%M%S")=='080000' or datetime.datetime.now().strftime("%H%M%S")=='140000' or datetime.datetime.now().strftime("%H%M%S")=='200000':
        for document in Patient_table.find({'Prescribed Medicines':{'$ne':None}}):
            for medicine  in document['Prescribed Medicines'].split('.'):
                old=Stock_table.find_one({'Medicine Name':medicine})
                new={'Medicine Name':old['Medicine Name'],'Units Remaining':int(old['Units Remaining'])-1}
                Stock_table.find_one_and_replace(old,new)
                print('Hi')
    for document in Stock_table.find({'Units Remaining':{'$lt':10}}):
        driver=webdriver.Chrome()
        driver.get('http://localhost:8000/index.html')
        driver.find_element_by_id(id_table.find_one({'Medicine Name':document['Medicine Name']})['id']).click()
        price+=medicine_price_table.find_one({'Medicine Name':document['Medicine Name']})['Price']
        mlist.append(document['Medicine Name'])
        old_dict=Stock_table.find_one({'Medicine Name': document['Medicine Name']})
        new_dict={'Medicine Name':document['Medicine Name'],'Units Remaining':document['Units Remaining']+ medicine_price_table.find_one({'Medicine Name':document['Medicine Name']})['Package units']}
        Stock_table.find_one_and_replace(old_dict,new_dict)
        driver.find_element_by_id('mycart').click()
        driver.find_element_by_id('checkout').click()
        driver.quit()
    if price != 0:
        print("\t\t\tBILL\n")
        for things in mlist:
            print("{}\t\t{}\n".format(things,medicine_price_table.find_one({'Medicine Name':things})['Price']))
        print("TOTAL PRICE\n"+str(price))
        price=0
        mlist.clear()
